# scripts/show_results.py
# Zeigt Trainings-Diagnose + einen vollen Anti-Martingale-Backtest (inkl. Gebuehren) fuer das
# aktuell trainierte Barriere-Modell an. Analog zum show_results.py/.sh-Muster der anderen
# Bots -- oraclebot hat nur EIN Symbol/Modell (kein Multi-Coin/Genome-Discovery wie dnabot),
# daher deutlich einfacher: keine Coin/Timeframe-Auswahl noetig, nur drei Modi.
import argparse
import json
import logging
import math
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

import pandas as pd

from oraclebot.analysis.evaluation import build_trades, run_anti_martingale_backtest
from oraclebot.data.dataset import load_dataset_jsonl
from oraclebot.model.barrier_model import BarrierPredictor
from oraclebot.utils.config import load_barrier_config, load_settings

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), '..')
ARTIFACTS_DIR = os.path.join(PROJECT_ROOT, 'artifacts', 'datasets')
CHARTS_DIR = os.path.join(PROJECT_ROOT, 'artifacts', 'charts')


def load_secrets() -> dict:
    secret_path = os.path.join(PROJECT_ROOT, 'secret.json')
    if not os.path.exists(secret_path):
        return {}
    with open(secret_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_predictions(barrier_cfg: dict):
    """Laedt Modell + gecachten Trainings-Datensatz, gibt die Modell-Vorhersagen fuer den
    Out-of-Sample-Validierungsanteil zurueck (chronologisch letzte val_split% der Beispiele --
    exakt derselbe Anteil, der beim Training als Holdout diente)."""
    symbol = barrier_cfg['symbol']
    reference_tf = barrier_cfg['reference_timeframe']
    safe_symbol = symbol.replace('/', '_').replace(':', '_')
    model_path = os.path.join(ARTIFACTS_DIR, f'barrier_model_{safe_symbol}_{reference_tf}.pkl')
    dataset_path = os.path.join(ARTIFACTS_DIR, f'barrier_{safe_symbol}_{reference_tf}.jsonl')
    if not os.path.exists(model_path):
        logger.error(f"Kein Modell gefunden: {model_path}. Erst train_barrier_model.py "
                     f"(bzw. run_pipeline.sh) ausfuehren.")
        sys.exit(1)
    if not os.path.exists(dataset_path):
        logger.error(f"Kein Trainings-Datensatz gefunden: {dataset_path}. Erst "
                     f"train_barrier_model.py ausfuehren (Datensatz-Cache ist nicht in Git).")
        sys.exit(1)

    predictor = BarrierPredictor.load(model_path)
    examples = load_dataset_jsonl(dataset_path)
    examples.sort(key=lambda e: e['date'])
    n_val = max(1, int(len(examples) * barrier_cfg['val_split']))
    val_examples = examples[-n_val:]

    preds = []
    for ex in val_examples:
        cls, conf = predictor.predict_one(ex['features'])
        preds.append({'date': ex['date'], 'entry': ex['entry'], 'cls': cls, 'conf': conf,
                      'label': ex['target'], 'exit_time': pd.Timestamp(ex['exit_time'])})
    preds.sort(key=lambda p: p['date'])
    return preds, safe_symbol


def print_summary(trades: list, backtest: dict, safe_symbol: str, reference_tf: str):
    wins = sum(1 for t in trades if t['outcome'] == 'win')
    win_rate = wins / len(trades) * 100 if trades else 0.0
    pnl_pct = (backtest['end_capital'] - backtest['start_capital']) / backtest['start_capital'] * 100

    diagnostics_path = os.path.join(ARTIFACTS_DIR, f'barrier_diagnostics_{safe_symbol}_{reference_tf}.json')
    logger.info('=' * 70)
    logger.info('TRAININGS-DIAGNOSE (letzter train_barrier_model.py-Lauf)')
    logger.info('=' * 70)
    if os.path.exists(diagnostics_path):
        with open(diagnostics_path, 'r', encoding='utf-8') as f:
            diag = json.load(f)
        logger.info(f"Beispiele gesamt: {diag['n_examples']} (Train={diag['n_train']} Val={diag['n_val']})")
        logger.info(f"70/30-Split: In-Sample={diag['train_accuracy']:.1%} Out-of-Sample={diag['val_accuracy']:.1%}")
        wf = diag['walk_forward_accuracies']
        logger.info(f"Walk-Forward ({len(wf)} Fenster): {['%.1f%%' % (a*100) for a in wf]}")
        logger.info(f"  Mittel={diag['walk_forward_mean']:.1%} Worst-Case={diag['walk_forward_worst_case']:.1%}")
    else:
        logger.info(f"Keine Diagnose-Datei gefunden ({diagnostics_path}).")

    logger.info('')
    logger.info('=' * 70)
    logger.info('BACKTEST MIT ANTI-MARTINGALE (Out-of-Sample, inkl. Gebuehren)')
    logger.info('=' * 70)
    logger.info(f"Trades={len(trades)} WinRate={win_rate:.1f}%")
    logger.info(f"Startkapital={backtest['start_capital']:.2f} USDT -> Endkapital={backtest['end_capital']:.2e} USDT "
                f"({'+' if pnl_pct >= 0 else ''}{pnl_pct:.1f}%)")
    logger.info(f"MaxDD={backtest['max_dd_pct']:.1f}%")
    logger.info("Hinweis: Endkapital bei vielen Trades + Compounding wird schnell astronomisch --")
    logger.info("nur als relativer Vergleichswert zwischen Konfigurationen aussagekraeftig, siehe README.")


def generate_chart(trades: list, barrier_cfg: dict, backtest: dict, safe_symbol: str, since: str = None):
    """Interaktives HTML-Chart (Plotly), Stil analog zu zerobots
    run_portfolio_optimizer.py::generate_equity_html -- Rangeslider, 'x unified'-Hover,
    plotly_white-Template, dieselben Win/Loss-Marker-Farben. Ersetzt das fruehere statische
    PNG (matplotlib) komplett, siehe README.

    `since`: wie bei generate_excel() -- filtert auf Trades ab diesem Datum UND laesst den
    Anti-Martingale-Backtest fuer die Kapitalkurve/MaxDD frisch bei backtest_start_capital neu
    laufen (nicht den bereits verzinsten Stand aus der Zeit vor `since` uebernehmen, siehe
    Bugfix 2026-07-26 in generate_excel())."""
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        logger.error("plotly nicht installiert -- Chart uebersprungen (pip install plotly).")
        return None

    from oraclebot.utils.data_fetch import fetch_all_timeframes

    if since:
        since_ts = pd.Timestamp(since, tz='UTC')
        trades = [t for t in trades if t['entry_time'] >= since_ts]
        if not trades:
            logger.error(f"Keine Trades ab {since} im Out-of-Sample-Zeitraum.")
            return None
        backtest = run_anti_martingale_backtest(trades, barrier_cfg)

    reference_tf = barrier_cfg['reference_timeframe']
    symbol = barrier_cfg['symbol']
    start_capital = barrier_cfg.get('backtest_start_capital', 15.0)

    ohlcv = fetch_all_timeframes(symbol, [reference_tf], barrier_cfg['history_days'],
                                  cache_dir=ARTIFACTS_DIR, use_cache=True)
    price_df = ohlcv[reference_tf]

    wins = [t for t in trades if t['outcome'] == 'win']
    liqs = [t for t in trades if t['outcome'] == 'loss']
    win_rate = len(wins) / len(trades) * 100

    # Kapitalkurve war bereits Teil von run_anti_martingale_backtest() (equity_after pro Trade).
    dates_ = [trades[0]['entry_time']] + [t['entry_time'] for t in trades]
    capitals = [start_capital] + [t['equity_after'] for t in trades]

    streak_values = []
    current_streak, current_type = 0, None
    for t in trades:
        if t['outcome'] == current_type:
            current_streak += 1
        else:
            current_type = t['outcome']
            current_streak = 1
        streak_values.append(current_streak if t['outcome'] == 'win' else -current_streak)
    max_win_streak = max(streak_values)
    max_liq_streak = abs(min(streak_values))

    plot_start = trades[0]['entry_time'] - pd.Timedelta(days=3)
    plot_end = trades[-1]['entry_time'] + pd.Timedelta(days=3)
    price_zoom = price_df[(price_df.index >= plot_start) & (price_df.index <= plot_end)]

    title = (f"{symbol} -- {reference_tf}-Barriere-Modell, conf>={barrier_cfg.get('min_confidence', 0.60):.2f} "
             f"(SL=TP={barrier_cfg['barrier_pct']:.0f}%, Hebel={barrier_cfg.get('leverage', 10.0)}x) | "
             f"{len(trades)} Trades | WR: {win_rate:.1f}% | MaxDD: {backtest['max_dd_pct']:.1f}%")

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.06,
                         row_heights=[0.5, 0.32, 0.18],
                         subplot_titles=(f"{symbol.split('/')[0]} Close ({reference_tf}) + Trades",
                                         f"Anti-Martingale-Kapitalkurve (Start={start_capital:.0f} USDT, inkl. Gebuehren)",
                                         f"Laufende Serie -- laengste Gewinn-Serie={max_win_streak}, "
                                         f"laengste Liq-Serie={max_liq_streak}"))

    fig.add_trace(go.Scatter(
        x=price_zoom.index, y=price_zoom['close'], mode='lines', name=f"{symbol.split('/')[0]} Close",
        line=dict(color='#94a3b8', width=1), opacity=0.8,
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=[t['entry_time'] for t in wins], y=[t['entry'] for t in wins], mode='markers',
        name=f'Win (n={len(wins)})',
        marker=dict(color='#26a69a', symbol='triangle-up', size=9, line=dict(width=1, color='#0f766e')),
    ), row=1, col=1)
    fig.add_trace(go.Scatter(
        x=[t['entry_time'] for t in liqs], y=[t['entry'] for t in liqs], mode='markers',
        name=f'Liquidation (n={len(liqs)})',
        marker=dict(color='#ef5350', symbol='triangle-down', size=9, line=dict(width=1, color='#7f1d1d')),
    ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=dates_, y=capitals, mode='lines', name=f"Anti-Martingale (Basis={barrier_cfg.get('anti_martingale_base_pct', 5.0):.2f}%)",
        line=dict(color='#2563eb', width=2), opacity=0.85,
    ), row=2, col=1)
    fig.add_hline(y=start_capital, line=dict(color='rgba(100,100,100,0.4)', width=1, dash='dash'),
                  annotation_text=f'Start {start_capital:.0f} USDT', annotation_position='top left', row=2, col=1)
    # Explizite Log-Range statt Plotly-Autorange: eine anti_martingale_backtest()-Kapitalkurve
    # kann bei einem katastrophalen Trade auf exakt 0.0 fallen (siehe evaluation.py's
    # `capital = max(capital, 0.0)`-Floor) -- ein einzelner 0-Wert in den Log-Achsen-Daten liess
    # Plotly beobachtet auf einen absurd riesigen Default-Bereich (bis 10^16) zurueckfallen,
    # obwohl die eigentlichen Werte nur ueber wenige Dekaden liegen (Nutzer-Feedback 2026-07-26).
    positive_capitals = [c for c in capitals if c > 0]
    if positive_capitals:
        log_min = math.log10(min(positive_capitals))
        log_max = math.log10(max(positive_capitals))
        pad = max((log_max - log_min) * 0.1, 0.15)
        fig.update_yaxes(type='log', title_text='USDT (log)', range=[log_min - pad, log_max + pad], row=2, col=1)
    else:
        fig.update_yaxes(type='log', title_text='USDT (log)', row=2, col=1)

    streak_colors = ['#26a69a' if v > 0 else '#ef5350' for v in streak_values]
    fig.add_trace(go.Bar(
        x=[t['entry_time'] for t in trades], y=streak_values, name='Serienlaenge',
        marker_color=streak_colors, showlegend=False,
    ), row=3, col=1)
    fig.add_hline(y=0, line=dict(color='black', width=0.8), row=3, col=1)
    fig.update_yaxes(title_text='Serienlaenge', row=3, col=1)

    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=900,
        hovermode='x unified',
        template='plotly_white',
        dragmode='zoom',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
        xaxis3=dict(rangeslider=dict(visible=True), range=[plot_start, plot_end]),
        xaxis=dict(range=[plot_start, plot_end]),
        xaxis2=dict(range=[plot_start, plot_end]),
    )
    fig.update_yaxes(title_text='USDT', row=1, col=1)

    os.makedirs(CHARTS_DIR, exist_ok=True)
    outfile = os.path.join(CHARTS_DIR, 'combined_overview.html')
    fig.write_html(outfile)
    logger.info(f"Chart gespeichert: {outfile} (laengste Gewinn-Serie={max_win_streak}, laengste Liq-Serie={max_liq_streak})")
    return outfile


def generate_excel(trades: list, barrier_cfg: dict, since: str = None):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl nicht installiert -- Excel-Export uebersprungen (pip install openpyxl).")
        return None

    if since:
        since_ts = pd.Timestamp(since, tz='UTC')
        trades = [t for t in trades if t['entry_time'] >= since_ts]
        if not trades:
            logger.error(f"Keine Trades ab {since} im Out-of-Sample-Zeitraum.")
            return None
        # Bugfix 2026-07-26: trades tragen an dieser Stelle noch margin_used/pnl_usdt/
        # equity_after aus dem VOLLEN, ungefilterten Anti-Martingale-Lauf oben (vor dem Filtern)
        # -- die Kapitalkurve wuerde also den bereits verzinsten Stand von VOR `since` einfach
        # fortsetzen statt wirklich frisch bei backtest_start_capital zu beginnen, obwohl genau
        # das dokumentiert ist ("Kapitalkurve startet fuer den Export dann frisch"). Deshalb den
        # Anti-Martingale-Backtest hier NOCHMAL, nur auf den gefilterten Trades, laufen lassen --
        # das ueberschreibt margin_used/pnl_usdt/equity_after mit Werten, die tatsaechlich bei
        # backtest_start_capital ab `since` neu anfangen.
        run_anti_martingale_backtest(trades, barrier_cfg)

    coin = barrier_cfg['symbol'].split('/')[0]
    reference_tf = barrier_cfg['reference_timeframe']
    rows = []
    for i, t in enumerate(trades, 1):
        outcome_label = 'TP erreicht' if t['outcome'] == 'win' else 'SL erreicht'
        rows.append({
            'Nr': i,
            'Datum (Entry)': t['entry_time'].strftime('%Y-%m-%d %H:%M'),
            'Datum (Exit)': t['exit_time'].strftime('%Y-%m-%d %H:%M'),
            'Coin': coin,
            'Timeframe': reference_tf,
            'Richtung': t['direction'].upper(),
            'Entry': round(t['entry'], 2),
            'Exit': round(t['exit'], 2),
            'Ergebnis': outcome_label,
            'Hebel': f"{t['leverage']:.0f}x",
            'Margin (USDT)': round(t['margin_used'], 4),
            'PnL (USDT)': round(t['pnl_usdt'], 4),
            'Gesamtkapital': round(t['equity_after'], 4),
        })

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    win_fill = PatternFill('solid', fgColor='D6F4DC')
    loss_fill = PatternFill('solid', fgColor='FAD7D7')
    alt_fill = PatternFill('solid', fgColor='F2F2F2')
    thin_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                          top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
    col_widths = {'Nr': 6, 'Datum (Entry)': 18, 'Datum (Exit)': 18, 'Coin': 10, 'Timeframe': 12,
                  'Richtung': 10, 'Entry': 14, 'Exit': 14, 'Ergebnis': 14, 'Hebel': 10,
                  'Margin (USDT)': 16, 'PnL (USDT)': 14, 'Gesamtkapital': 16}

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        fill = win_fill if row['Ergebnis'] == 'TP erreicht' else (loss_fill if r_idx % 2 == 0 else alt_fill)
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'Margin (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    start_capital = barrier_cfg.get('backtest_start_capital', 15.0)
    end_capital = rows[-1]['Gesamtkapital']
    pnl_pct = (end_capital - start_capital) / start_capital * 100
    win_rate = sum(1 for t in trades if t['outcome'] == 'win') / len(trades) * 100
    sign = '+' if pnl_pct >= 0 else ''

    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    summary_row += 1
    am_base = barrier_cfg.get('anti_martingale_base_pct', 5.0)
    am_growth = barrier_cfg.get('anti_martingale_growth_factor', 2.0)
    am_streak = barrier_cfg.get('anti_martingale_streak_target', 3)
    fee_pct = barrier_cfg.get('taker_fee_rate_pct', 0.06)
    for label, value in [
        ('Trades gesamt', len(rows)),
        ('Win-Rate', f'{win_rate:.1f}%'),
        ('PnL', f'{sign}{pnl_pct:.1f}%'),
        ('Final Equity', f'{end_capital:.2f} USDT'),
        ('Startkapital', f'{start_capital:.2f} USDT'),
        ('Anti-Martingale', f'Basis={am_base:.2f}% Growth={am_growth:.1f}x Streak-Ziel={am_streak}'),
        ('Gebuehren', f'{fee_pct:.2f}%/Seite Taker (Roundtrip), bereits in PnL/Margin eingerechnet'),
        ('Zeitraum', f"{trades[0]['entry_time'].date()} -> {trades[-1]['exit_time'].date()}"),
        ('Hinweis', 'Backtest auf Out-of-Sample-Validierungsdaten, KEIN Live-Trade-Log.'),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    ts_stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(CHARTS_DIR, exist_ok=True)
    outfile = os.path.join(CHARTS_DIR, f'oraclebot_trades_{ts_stamp}.xlsx')
    wb.save(outfile)
    logger.info(f"Excel-Tabelle gespeichert: {outfile}")
    return outfile


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--chart', action='store_true', help="Aktualisiert artifacts/charts/combined_overview.png.")
    parser.add_argument('--excel', action='store_true', help="Erstellt artifacts/charts/oraclebot_trades_<Zeitstempel>.xlsx.")
    parser.add_argument('--since', type=str, default=None,
                         help="Nur Trades ab diesem Datum (JJJJ-MM-TT) fuer --chart/--excel beruecksichtigen. "
                              "Kapitalkurve startet dann frisch bei backtest_start_capital.")
    parser.add_argument('--start-capital', type=float, default=None,
                         help="Ueberschreibt barrier_strategy_settings.backtest_start_capital fuer diesen Lauf "
                              "(betrifft Zusammenfassung, Chart und Excel-Export gleichermassen).")
    args = parser.parse_args()

    settings = load_settings()
    barrier_cfg = load_barrier_config(settings)
    if args.start_capital is not None:
        barrier_cfg['backtest_start_capital'] = args.start_capital
    preds, safe_symbol = load_predictions(barrier_cfg)
    trades = build_trades(preds, barrier_cfg)
    if not trades:
        logger.error("Keine Trades im Out-of-Sample-Zeitraum (min_confidence evtl. zu hoch?).")
        sys.exit(1)

    backtest = run_anti_martingale_backtest(trades, barrier_cfg)
    print_summary(trades, backtest, safe_symbol, barrier_cfg['reference_timeframe'])
    if args.since:
        logger.info(f"Hinweis: Obige Zusammenfassung bezieht sich auf den GESAMTEN "
                    f"Out-of-Sample-Zeitraum -- der --since-Filter ({args.since}) gilt nur fuer "
                    f"Chart/Excel-Export unten, die separat und mit frisch bei "
                    f"backtest_start_capital neu startender Kapitalkurve berechnet werden.")

    telegram_enabled = settings.get('notification_settings', {}).get('telegram_enabled', False)
    telegram_cfg = load_secrets().get('telegram', {}) if telegram_enabled else {}
    win_rate = sum(1 for t in trades if t['outcome'] == 'win') / len(trades) * 100

    if args.chart:
        logger.info('')
        chart_path = generate_chart(trades, barrier_cfg, backtest, safe_symbol, since=args.since)
        if chart_path and telegram_enabled:
            # HTML (interaktives Plotly-Chart) -- Telegram wuerde eine .html-Datei per
            # sendPhoto entweder ablehnen oder als reine Bilddatei fehlinterpretieren, daher
            # sendDocument (wie beim Excel-Export), auch wenn die Datei nicht direkt in
            # Telegram angezeigt wird -- lokal herunterladen und im Browser oeffnen.
            from oraclebot.utils.telegram import send_document
            chart_trades, chart_backtest = trades, backtest
            if args.since:
                since_ts = pd.Timestamp(args.since, tz='UTC')
                chart_trades = [t for t in trades if t['entry_time'] >= since_ts]
                chart_backtest = run_anti_martingale_backtest(chart_trades, barrier_cfg)
            chart_win_rate = sum(1 for t in chart_trades if t['outcome'] == 'win') / len(chart_trades) * 100
            since_note = f" (ab {args.since})" if args.since else ""
            caption = (f"oraclebot combined_overview.html{since_note}\n"
                       f"{len(chart_trades)} Trades, {chart_win_rate:.1f}% Winrate, "
                       f"MaxDD {chart_backtest['max_dd_pct']:.1f}%")
            send_document(telegram_cfg.get('bot_token'), telegram_cfg.get('chat_id'), chart_path, caption=caption)

    if args.excel:
        logger.info('')
        excel_path = generate_excel(trades, barrier_cfg, since=args.since)
        if excel_path and telegram_enabled:
            from oraclebot.utils.telegram import send_document
            excel_trades = trades
            if args.since:
                since_ts = pd.Timestamp(args.since, tz='UTC')
                excel_trades = [t for t in trades if t['entry_time'] >= since_ts]
            excel_win_rate = sum(1 for t in excel_trades if t['outcome'] == 'win') / len(excel_trades) * 100
            since_note = f" (ab {args.since})" if args.since else ""
            caption = f"oraclebot Trade-Log-Export{since_note}: {len(excel_trades)} Trades, {excel_win_rate:.1f}% Winrate"
            send_document(telegram_cfg.get('bot_token'), telegram_cfg.get('chat_id'), excel_path, caption=caption)
