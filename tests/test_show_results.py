import re

import openpyxl
import pandas as pd
import pytest

import oraclebot.utils.data_fetch as data_fetch_mod
import scripts.show_results as sr
from oraclebot.analysis.evaluation import run_anti_martingale_backtest

BASE_CFG = {'barrier_pct': 1.0, 'leverage': 100, 'backtest_start_capital': 15.0,
            'taker_fee_rate_pct': 0.06, 'symbol': 'BTC/USDT:USDT', 'reference_timeframe': '4h',
            'anti_martingale_base_pct': 5.0, 'anti_martingale_growth_factor': 2.0,
            'anti_martingale_streak_target': 3, 'history_days': 100}


def make_trade(entry_time, frac, outcome):
    ts = pd.Timestamp(entry_time, tz='UTC')
    return {'entry_time': ts, 'exit_time': ts, 'entry': 60000.0, 'exit': 60000.0,
            'direction': 'long', 'frac': frac, 'outcome': outcome}


@pytest.fixture
def isolated_charts_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(sr, 'CHARTS_DIR', str(tmp_path))
    return tmp_path


def test_since_filter_restarts_capital_curve_fresh_instead_of_carrying_over_prior_compounding(
        isolated_charts_dir):
    # Vier Gewinn-Trades, die ersten zwei VOR dem since-Datum -- ein voller (ungefilterter)
    # Anti-Martingale-Lauf laesst das Kapital vor dem since-Datum bereits deutlich wachsen.
    trades = [
        make_trade('2026-01-01', 0.05, 'win'),
        make_trade('2026-01-02', 0.05, 'win'),
        make_trade('2026-02-01', 0.05, 'win'),
        make_trade('2026-02-02', 0.05, 'win'),
    ]
    run_anti_martingale_backtest(trades, BASE_CFG)  # simuliert den vollen OOS-Lauf in show_results.py
    inflated_third_trade_equity = trades[2]['equity_after']

    outfile = sr.generate_excel(trades, BASE_CFG, since='2026-02-01')
    assert outfile is not None

    # Bugfix-Kern: eine unabhaengige, frische Simulation NUR der beiden gefilterten Trades muss
    # exakt denselben Endstand ergeben wie generate_excel() intern berechnet -- keine
    # Uebernahme des bereits (durch die ersten beiden Trades) aufgezinsten Zwischenstands.
    fresh_trades = [make_trade('2026-02-01', 0.05, 'win'), make_trade('2026-02-02', 0.05, 'win')]
    run_anti_martingale_backtest(fresh_trades, BASE_CFG)

    wb = openpyxl.load_workbook(outfile)
    ws = wb.active
    header = [c.value for c in ws[1]]
    capital_col = header.index('Gesamtkapital') + 1
    written_capitals = [ws.cell(row=r, column=capital_col).value for r in (2, 3)]

    # abs statt rel-Toleranz: generate_excel() rundet 'Gesamtkapital' bewusst auf 4 Dezimalstellen
    # fuer die Anzeige (round(t['equity_after'], 4)), eine rel=1e-6-Toleranz ist bei groesseren
    # Werten enger als diese Rundung selbst erlauben kann.
    assert written_capitals == pytest.approx([fresh_trades[0]['equity_after'], fresh_trades[1]['equity_after']],
                                              abs=1e-4)
    # Vor dem Fix waere hier der (durch die ersten beiden Trades) bereits aufgezinste Stand
    # gelandet -- deutlich hoeher als der frische Wert.
    assert written_capitals[0] < inflated_third_trade_equity


def test_since_filter_with_no_matching_trades_returns_none(isolated_charts_dir):
    trades = [make_trade('2026-01-01', 0.05, 'win')]
    run_anti_martingale_backtest(trades, BASE_CFG)
    assert sr.generate_excel(trades, BASE_CFG, since='2027-01-01') is None


def test_without_since_filter_keeps_full_period_capital_curve(isolated_charts_dir):
    trades = [make_trade('2026-01-01', 0.05, 'win'), make_trade('2026-01-02', 0.05, 'win')]
    run_anti_martingale_backtest(trades, BASE_CFG)
    expected_final_equity = trades[-1]['equity_after']

    outfile = sr.generate_excel(trades, BASE_CFG, since=None)
    wb = openpyxl.load_workbook(outfile)
    ws = wb.active
    header = [c.value for c in ws[1]]
    capital_col = header.index('Gesamtkapital') + 1
    assert ws.cell(row=3, column=capital_col).value == pytest.approx(expected_final_equity, rel=1e-6)


@pytest.fixture
def fake_price_data(monkeypatch):
    def fake_fetch_all_timeframes(symbol, timeframes, history_days, cache_dir=None, use_cache=True):
        idx = pd.date_range('2026-01-01', periods=500, freq='4h', tz='UTC')
        df = pd.DataFrame({'open': 60000.0, 'high': 60100.0, 'low': 59900.0, 'close': 60000.0,
                            'volume': 100.0}, index=idx)
        return {timeframes[0]: df}
    monkeypatch.setattr(data_fetch_mod, 'fetch_all_timeframes', fake_fetch_all_timeframes)


def test_chart_since_filter_also_restarts_capital_curve_fresh(isolated_charts_dir, fake_price_data):
    # Derselbe Bug wie bei generate_excel() (Bugfix 2026-07-26): generate_chart() bekam --since
    # zunaechst gar nicht erst uebergeben und haette denselben aufgezinsten Zwischenstand
    # weitergefuehrt statt frisch neu zu starten.
    trades = [
        make_trade('2026-01-01', 0.05, 'win'),
        make_trade('2026-01-02', 0.05, 'win'),
        make_trade('2026-02-01', 0.05, 'win'),
        make_trade('2026-02-02', 0.05, 'win'),
    ]
    backtest = run_anti_martingale_backtest(trades, BASE_CFG)
    inflated_third_trade_equity = trades[2]['equity_after']

    outfile = sr.generate_chart(trades, BASE_CFG, backtest, 'BTC_USDT_USDT', since='2026-02-01')
    assert outfile is not None

    fresh_trades = [make_trade('2026-02-01', 0.05, 'win'), make_trade('2026-02-02', 0.05, 'win')]
    run_anti_martingale_backtest(fresh_trades, BASE_CFG)

    # generate_chart() filtert lokal (neue Liste), mutiert aber dieselben Trade-Dict-Objekte per
    # run_anti_martingale_backtest() -- trades[2]/trades[3] muessen also jetzt den frischen,
    # nicht mehr den aufgezinsten Stand tragen.
    assert trades[2]['equity_after'] == pytest.approx(fresh_trades[0]['equity_after'], rel=1e-6)
    assert trades[3]['equity_after'] == pytest.approx(fresh_trades[1]['equity_after'], rel=1e-6)
    assert trades[2]['equity_after'] < inflated_third_trade_equity


def test_chart_since_filter_with_no_matching_trades_returns_none(isolated_charts_dir, fake_price_data):
    trades = [make_trade('2026-01-01', 0.05, 'win')]
    backtest = run_anti_martingale_backtest(trades, BASE_CFG)
    assert sr.generate_chart(trades, BASE_CFG, backtest, 'BTC_USDT_USDT', since='2027-01-01') is None


def test_chart_yaxis_range_stays_sane_even_after_a_total_wipeout_to_zero_capital(
        isolated_charts_dir, fake_price_data):
    # Ein katastrophaler Trade (frac=-1.0 bei 100x Hebel) drueckt das Kapital auf exakt 0.0
    # (siehe evaluation.py's `capital = max(capital, 0.0)`-Floor) -- ein einzelner 0-Wert in den
    # Log-Achsen-Daten liess Plotlys Autorange beobachtet auf einen absurd riesigen
    # Default-Bereich (bis 10^16) zurueckfallen, obwohl die eigentlichen Werte nur ueber wenige
    # Dekaden liegen (Nutzer-Feedback 2026-07-26).
    trades = [
        make_trade('2026-01-01', 0.05, 'win'),
        make_trade('2026-01-02', -1.0, 'loss'),
        make_trade('2026-01-03', 0.05, 'win'),
        make_trade('2026-01-04', 0.05, 'win'),
    ]
    backtest = run_anti_martingale_backtest(trades, BASE_CFG)
    assert 0.0 in [t['equity_after'] for t in trades]  # Testannahme bestaetigen

    outfile = sr.generate_chart(trades, BASE_CFG, backtest, 'BTC_USDT_USDT')
    html = open(outfile, encoding='utf-8').read()
    # Nicht-gierig ueber verschachtelte "title":{...}-Klammern hinweg bis zum naechsten "range".
    m = re.search(r'"yaxis2":\{.*?"range":\[([\-0-9.]+),\s*([\-0-9.]+)\]', html)
    assert m is not None, "yaxis2 range nicht im HTML gefunden"
    log_lo, log_hi = float(m.group(1)), float(m.group(2))
    # Vor dem Fix haette Plotlys Autorange hier problemlos > 10 Dekaden (z.B. bis 10^16)
    # aufspannen koennen -- mit expliziter Range darf die Spanne nur wenige Dekaden betragen.
    assert log_hi - log_lo < 5
